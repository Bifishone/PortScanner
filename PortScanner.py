import socket
import argparse
import sys
import time
import threading
import os
from queue import Queue
from typing import List, Tuple, Dict
from colorama import Fore, init, Style
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# 初始化colorama，确保跨平台彩色输出正常工作
init(autoreset=True)

# 线程锁，确保打印输出和数据操作不会混乱
print_lock = threading.Lock()
data_lock = threading.Lock()


def get_unique_filename(base_name: str, extension: str) -> str:
    """生成不重复的文件名，避免覆盖已有文件"""
    counter = 1
    filename = f"{base_name}.{extension}"

    while os.path.exists(filename):
        filename = f"{base_name}_{counter}.{extension}"
        counter += 1

    return filename


def export_to_excel(results: List[Dict]) -> None:
    """将扫描结果导出到Excel文件并进行美化处理"""
    if not results:
        print(f"{Fore.YELLOW}没有开放的端口，不生成Excel文件")
        return

    # 获取不重复的文件名
    filename = get_unique_filename("result", "xlsx")

    # 创建工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "开放端口扫描结果"

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    open_status_font = Font(color="008000", bold=True)  # 开放状态绿色加粗

    # 设置表头
    headers = ["target", "ip", "port", "PortIntroduction", "status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 填充数据
    for row, result in enumerate(results, 2):
        # 偶数行添加灰色背景，提高可读性
        row_fill = even_row_fill if row % 2 == 0 else None

        # 填充各列数据
        cell = ws.cell(row=row, column=1)
        cell.value = result["target"]
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if row_fill:
            cell.fill = row_fill

        cell = ws.cell(row=row, column=2)
        cell.value = result["ip"]
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if row_fill:
            cell.fill = row_fill

        cell = ws.cell(row=row, column=3)
        cell.value = result["port"]
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if row_fill:
            cell.fill = row_fill

        cell = ws.cell(row=row, column=4)
        cell.value = result["PortIntroduction"]
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        if row_fill:
            cell.fill = row_fill
        # 未知端口标灰
        if result["PortIntroduction"] == "Unknown":
            cell.font = Font(color="808080")

        cell = ws.cell(row=row, column=5)
        cell.value = result["status"]
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = open_status_font
        if row_fill:
            cell.fill = row_fill

    # 自动调整列宽
    for col in range(1, len(headers) + 1):
        max_length = 0
        # 检查该列所有单元格
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # 设置列宽（留一些余量）
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col)].width = adjusted_width

    # 冻结表头，方便滚动查看
    ws.freeze_panes = "A2"

    # 保存文件
    wb.save(filename)
    print(f"\n{Fore.GREEN}扫描结果已导出到Excel文件: {filename}")


def print_aligned_banner():
    """打印格式化的彩色Banner和作者信息"""
    # 清屏（适配Windows/Linux/macOS）
    os.system('cls' if os.name == 'nt' else 'clear')

    # 定义PortScanner的ASCII图案行
    banner_lines = [
        "____            _   ____",
        "|  _ \\ ___  _ __| |_/ ___|  ___ __ _ _ __  _ __   ___ _ __",
        "| |_) / _ \\| '__| __\\___ \\ / __/ _` | '_ \\| '_ \\ / _ \\ '__|",
        "|  __/ (_) | |  | |_ ___) | (_| (_| | | | | | | |  __/ |",
        "|_|   \\___/|_|   \\__|____/ \\___\\__,_|_| |_|_| |_|\\___|_|"
    ]

    # 计算所有行的最大长度（用于统一对齐）
    max_line_length = max(len(line) for line in banner_lines)

    # 打印彩色对齐的Banner
    print("\n" * 2)  # 顶部留白
    for i, line in enumerate(banner_lines):
        # 为每行分配不同颜色（循环使用青、绿、黄、品红、蓝）
        colors = [Fore.CYAN, Fore.GREEN, Fore.YELLOW, Fore.MAGENTA, Fore.BLUE]
        colored_line = colors[i % len(colors)] + line.ljust(max_line_length)  # 左对齐并补全长度
        print(colored_line)

    # 打印分隔线与作者信息
    divider = "-" * (max_line_length + 10)
    print("\n" + Fore.GREEN + divider)
    author_info = [
        f"{Fore.CYAN}Author: {Fore.WHITE}Bifish",
        f"{Fore.CYAN}GitHub: {Fore.WHITE}https://github.com/Bifishone"
    ]
    for info in author_info:
        print(info.center(max_line_length + 10))  # 作者信息居中对齐
    print(Fore.YELLOW + divider + "\n" * 2)


def parse_ports(port_str: str) -> List[int]:
    """解析端口字符串，支持多种格式：单个端口、范围、多个端口"""
    ports = []
    # 处理多个端口的情况（用逗号分隔）
    if ',' in port_str:
        for part in port_str.split(','):
            ports.extend(parse_ports(part.strip()))
        return sorted(list(set(ports)))  # 去重并排序

    # 处理端口范围的情况（用短横线分隔）
    if '-' in port_str:
        start_end = port_str.split('-')
        if len(start_end) != 2:
            raise ValueError(f"无效的端口范围格式: {port_str}")

        try:
            start = int(start_end[0].strip())
            end = int(start_end[1].strip())
        except ValueError:
            raise ValueError(f"端口必须是整数: {port_str}")

        if start > end:
            start, end = end, start

        for port in range(start, end + 1):
            if 1 <= port <= 65535:
                ports.append(port)
        return ports

    # 处理单个端口的情况
    try:
        port = int(port_str.strip())
        if 1 <= port <= 65535:
            ports.append(port)
            return ports
        else:
            raise ValueError(f"端口必须在1到65535之间: {port}")
    except ValueError:
        raise ValueError(f"无效的端口格式: {port_str}")


def read_ports_from_file(filename: str) -> List[int]:
    """从文件中读取端口列表，支持带#注释的行和多种编码"""
    ports = []
    # 尝试多种常见编码格式
    encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16', 'utf-8-sig']

    for encoding in encodings:
        try:
            with open(filename, 'r', encoding=encoding) as f:
                for line in f:
                    # 分割注释部分，只处理#之前的内容
                    port_part = line.split('#', 1)[0].strip()
                    if port_part:  # 忽略空行和纯注释行
                        ports.extend(parse_ports(port_part))
            # 如果成功读取，去重排序后返回
            return sorted(list(set(ports)))
        except UnicodeDecodeError:
            continue  # 尝试下一种编码
        except FileNotFoundError:
            raise FileNotFoundError(f"端口文件不存在: {filename}")
        except Exception as e:
            raise Exception(f"读取端口文件时出错: {str(e)}")

    # 如果所有编码都尝试失败
    raise UnicodeDecodeError(f"无法解析文件 {filename}，尝试了以下编码: {', '.join(encodings)}")


def read_ips_from_file(filename: str) -> List[str]:
    """从文件中读取IP地址列表，支持带#注释的行和多种编码"""
    ips = []
    # 尝试多种常见编码格式
    encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16', 'utf-8-sig']

    for encoding in encodings:
        try:
            with open(filename, 'r', encoding=encoding) as f:
                for line in f:
                    # 分割注释部分，只处理#之前的内容
                    ip_part = line.split('#', 1)[0].strip()
                    if ip_part:  # 忽略空行和纯注释行
                        ips.append(ip_part)
            # 如果成功读取，去重后返回
            return list(set(ips))
        except UnicodeDecodeError:
            continue  # 尝试下一种编码
        except FileNotFoundError:
            raise FileNotFoundError(f"IP文件不存在: {filename}")
        except Exception as e:
            raise Exception(f"读取IP文件时出错: {str(e)}")

    # 如果所有编码都尝试失败
    raise UnicodeDecodeError(f"无法解析文件 {filename}，尝试了以下编码: {', '.join(encodings)}")


def load_port_descriptions(filename: str) -> Dict[int, str]:
    """从port.ini文件加载端口描述信息"""
    port_desc = {}
    encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16', 'utf-8-sig']

    for encoding in encodings:
        try:
            with open(filename, 'r', encoding=encoding) as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue

                    # 分割端口号和注释部分
                    parts = line.split('#', 1)
                    if len(parts) >= 2:
                        port_part = parts[0].strip()
                        desc_part = parts[1].strip()

                        try:
                            port = int(port_part)
                            if 1 <= port <= 65535:
                                port_desc[port] = desc_part
                        except ValueError:
                            continue  # 忽略非数字的端口行
            return port_desc
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            print(f"{Fore.YELLOW}警告: 端口描述文件 {filename} 不存在，所有端口将显示为Unknown")
            return {}
        except Exception as e:
            print(f"{Fore.YELLOW}警告: 读取端口描述文件时出错: {str(e)}，所有端口将显示为Unknown")
            return {}

    # 如果所有编码都尝试失败
    print(f"{Fore.YELLOW}警告: 无法解析端口描述文件 {filename}，所有端口将显示为Unknown")
    return {}


def check_port(ip: str, port: int, timeout: float = 3.0) -> Tuple[bool, str]:
    """检查指定IP的端口是否开放"""
    try:
        # 创建socket对象
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            # 设置超时时间
            sock.settimeout(timeout)

            # 尝试连接
            result = sock.connect_ex((ip, port))

            if result == 0:
                return (True, "开放")
            else:
                return (False, "关闭")

    except socket.timeout:
        return (False, f"超时({timeout}秒)")
    except socket.gaierror:
        return (False, "无效IP地址")
    except socket.error as e:
        return (False, f"错误: {str(e)}")


def scan_ips_ports(ips: List[str], ports: List[int], port_descriptions: Dict[int, str],
                   timeout: float = 3.0, threads: int = 5) -> List[Dict]:
    """使用多线程扫描多个IP和端口，返回扫描结果用于导出"""
    total_tasks = len(ips) * len(ports)
    if total_tasks == 0:
        print(f"{Fore.YELLOW}没有需要扫描的任务")
        return []

    # 创建任务队列
    task_queue = Queue()
    for ip in ips:
        for port in ports:
            task_queue.put((ip, port))

    # 用于存储开放端口的字典 {ip: [ports]}
    open_ports = {ip: [] for ip in ips}

    # 用于存储导出结果的数据列表
    export_results = []

    progress_counter = 0
    progress_lock = threading.Lock()

    # 打印扫描开始信息
    print(f"{Fore.YELLOW}{'-' * 80}")
    print(f"{Fore.WHITE}开始扫描: {len(ips)} 个IP, {len(ports)} 个端口")
    print(f"{Fore.WHITE}超时时间: {timeout} 秒, 线程数量: {threads}")
    print(f"{Fore.GREEN}{'-' * 80}\n")
    start_time = time.time()

    # 定义线程工作函数
    def scan_worker():
        nonlocal progress_counter
        while not task_queue.empty():
            try:
                # 非阻塞方式获取队列元素
                ip, port = task_queue.get_nowait()
            except:
                break

            try:
                is_open, status = check_port(ip, port, timeout)

                # 更新进度
                with progress_lock:
                    progress_counter += 1
                    current = progress_counter
                    progress = (current / total_tasks) * 100

                # 输出结果
                with print_lock:
                    # 显示进度
                    sys.stdout.write(f"\r{Fore.RED}扫描进度: {Fore.YELLOW}{progress:.1f}% ({current}/{total_tasks})")
                    sys.stdout.flush()

                    # 只显示开放的端口
                    if is_open:
                        # 获取端口描述，如果没有则为"Unknown"
                        port_desc = port_descriptions.get(port, "Unknown")
                        # 为Unknown描述设置灰色，其他使用青色
                        desc_color = Fore.CYAN if port_desc != "Unknown" else Fore.LIGHTBLACK_EX

                        # 格式化输出，三列严格对齐
                        print(f"\n{Fore.WHITE}{ip}:{port:<30} {desc_color}{port_desc:<40} {Fore.GREEN}{status:>20}")
                        open_ports[ip].append(port)

                        # 添加到导出结果列表（线程安全）
                        with data_lock:
                            export_results.append({
                                "target": f"{ip}:{port}",
                                "ip": ip,
                                "port": port,
                                "PortIntroduction": port_desc,
                                "status": status
                            })

            except Exception as e:
                with print_lock:
                    print(f"\n{Fore.RED}扫描 {ip}:{port} 时出错: {str(e)}")
            finally:
                task_queue.task_done()

    # 创建并启动线程
    thread_list = []
    for _ in range(threads):
        thread = threading.Thread(target=scan_worker)
        thread.daemon = True  # 守护线程
        thread_list.append(thread)
        thread.start()

    # 等待所有任务完成
    task_queue.join()

    # 确保所有线程都已结束
    for thread in thread_list:
        thread.join()

    end_time = time.time()
    elapsed = end_time - start_time

    # 统计总开放端口数并排序
    total_open = 0
    for ip in open_ports:
        total_open += len(open_ports[ip])
        open_ports[ip].sort()

    # 输出最终统计结果
    print(f"\n\n{Fore.GREEN}{'-' * 80}")
    print(f"{Fore.YELLOW}扫描完成! 耗时: {elapsed:.2f} 秒")
    print(f"{Fore.YELLOW}总开放端口数: {total_open}")

    # 输出每个IP的开放端口列表
    for ip in open_ports:
        if open_ports[ip]:
            print(f"\n{Fore.BLUE}IP: {ip} 开放的端口:")
            port_info = []
            for port in open_ports[ip]:
                desc = port_descriptions.get(port, "Unknown")
                # 为Unknown描述设置灰色
                desc_color = Fore.GREEN if desc != "Unknown" else Fore.LIGHTBLACK_EX
                port_info.append(f"{port:<5}({desc_color}{desc}{Fore.GREEN})")

            # 用逗号拼接成整齐的列表
            print(f"{Fore.GREEN}  " + ", ".join(port_info))

    print(f"\n{Fore.YELLOW}{'-' * 80}")

    return export_results


def main():
    # 显示banner
    print_aligned_banner()

    # 加载端口描述信息
    port_descriptions = load_port_descriptions("port.ini")

    # 设置命令行参数
    parser = argparse.ArgumentParser(description='多线程端口扫描工具，支持导出结果到Excel',
                                     formatter_class=argparse.RawTextHelpFormatter)

    # IP参数组（互斥）
    ip_group = parser.add_mutually_exclusive_group(required=True)
    ip_group.add_argument('-ip', help='指定单个IP地址')
    ip_group.add_argument('-ip-list', help='从文件中读取IP地址列表（支持#注释）')

    # 端口参数组（互斥）
    port_group = parser.add_mutually_exclusive_group(required=True)
    port_group.add_argument('-p', help='指定端口，可以是：\n'
                                       '  - 单个端口: -p 80\n'
                                       '  - 端口范围: -p 1-80\n'
                                       '  - 多个端口: -p 80,25,443')
    port_group.add_argument('-p-list', help='从文件中读取端口列表（支持#注释）')

    # 其他参数
    parser.add_argument('-t', '--timeout', type=float, default=3.0,
                        help='超时时间(秒)，默认3秒')
    parser.add_argument('-threads', type=int, default=5,
                        help='并行线程数量，默认5个')

    args = parser.parse_args()

    try:
        # 验证线程数量
        if args.threads < 1 or args.threads > 100:
            raise ValueError("线程数量必须在1到100之间")

        # 处理IP
        ips = []
        if args.ip:
            ips.append(args.ip)
        elif args.ip_list:
            ips = read_ips_from_file(args.ip_list)

        if not ips:
            print(f"{Fore.RED}错误: 没有有效的IP地址")
            return

        # 处理端口
        ports = []
        if args.p:
            ports = parse_ports(args.p)
        elif args.p_list:
            ports = read_ports_from_file(args.p_list)

        if not ports:
            print(f"{Fore.RED}错误: 没有有效的端口")
            return

        # 执行扫描，获取结果
        scan_results = scan_ips_ports(ips, ports, port_descriptions, args.timeout, args.threads)

        # 导出结果到Excel
        export_to_excel(scan_results)

    except Exception as e:
        print(f"{Fore.RED}错误: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    main()
